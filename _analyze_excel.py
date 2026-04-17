from openpyxl import load_workbook
from collections import Counter

xlsx = r"c:/Users/jaruani/Downloads/Python/App de potencia/GRAFICOS DE POTENCIAS 2026-MARZO.xlsx"
wb = load_workbook(xlsx, data_only=False, read_only=False)

print("=== SHEETS ===")
for ws in wb.worksheets:
    print(f"{ws.title}|dim={ws.calculate_dimension()}|max_row={ws.max_row}|max_col={ws.max_column}")

print("\n=== TABLES ===")
found = 0
for ws in wb.worksheets:
    for name, t in ws.tables.items():
        found += 1
        print(f"{ws.title}|{name}|{t.ref}")
if not found:
    print("NONE")

print("\n=== PIVOTS ===")
pcount = 0
for ws in wb.worksheets:
    pivots = getattr(ws, "_pivots", []) or []
    for p in pivots:
        pcount += 1
        n = getattr(p, "name", None) or getattr(p, "tableName", None) or "NO_NAME"
        loc = getattr(p, "location", None)
        pref = getattr(loc, "ref", None) if loc else None
        print(f"{ws.title}|{n}|{pref}")
if pcount == 0:
    print("NONE")

keywords = [
    "POT_DEMANDADA","POT_FACT_PUNTA","POT_FACT_VALLE","POT_FACT_LLANO",
    "MAX(","MAXIMO(","NIC","COMBINACION","SECTOR","BUSCARV(","XLOOKUP(",
    "INDICE(","COINCIDIR(","FILTRAR("
]

hits = []
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                u = v.upper()
                if any(k in u for k in keywords):
                    hits.append((ws.title, c.coordinate, v))

print("\n=== FORMULA_HITS_COUNT ===")
print(len(hits))
print("\n=== FORMULA_HITS_SAMPLE_20 ===")
for t,coord,f in hits[:20]:
    print(f"{t}|{coord}|{f}")

print("\n=== FORMULAS_MAX_POT_FACT ===")
max_hits = []
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                u = v.upper()
                if "MAX(" in u and ("POT_FACT_PUNTA" in u or "POT_FACT_VALLE" in u or "POT_FACT_LLANO" in u):
                    max_hits.append((ws.title, c.coordinate, v))
for t,coord,f in max_hits[:20]:
    print(f"{t}|{coord}|{f}")
print(f"TOTAL_MAX_FORMULAS|{len(max_hits)}")

print("\n=== NIC_REFERENCE_HEADER_SCAN ===")
for ws in wb.worksheets:
    for r in range(1, min(ws.max_row, 80)+1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 80)+1)]
        u = [str(v).strip().upper() if v is not None else "" for v in vals]
        score = 0
        score += 1 if any(x in {"N° NIC", "N NIC", "NIC"} for x in u) else 0
        score += 1 if "SECTOR" in u else 0
        score += 1 if "DENOMINACION" in u else 0
        score += 1 if "TARIFA" in u else 0
        score += 1 if "COMBINACION" in u else 0
        score += 1 if "SECTOR MACRO" in u else 0
        if score >= 4:
            print(f"{ws.title}|row={r}|score={score}")
            cells = []
            for c in range(1, min(ws.max_column, 80)+1):
                vv = ws.cell(r, c).value
                if vv not in (None, ""):
                    cells.append(f"{c}:{vv}")
            print("HEADERS|" + " ; ".join(cells[:20]))
            break
